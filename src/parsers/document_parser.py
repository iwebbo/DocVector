# src/parsers/document_parser.py
from .base_parser import BaseParser
from pathlib import Path
from typing import List, Dict
import docx
import PyPDF2
import openpyxl

class DocumentParser(BaseParser):
    def parse(self, file_path: Path) -> List[Dict]:
        ext = file_path.suffix.lower()
        
        if ext == '.pdf':
            return self._parse_pdf(file_path)
        elif ext in ['.docx', '.doc']:
            return self._parse_docx(file_path)
        elif ext in ['.xlsx', '.xls']:
            return self._parse_excel(file_path)
        else:
            return self._parse_text(file_path)
    
    def _parse_pdf(self, file_path: Path) -> List[Dict]:
        chunks = []
        with open(file_path, 'rb') as f:
            reader = PyPDF2.PdfReader(f)
            for page_num, page in enumerate(reader.pages):
                text = page.extract_text()
                if text.strip():
                    chunks.append({
                        "text": text,
                        "metadata": {
                            **self.get_file_metadata(file_path),
                            "source_type": "document",
                            "page": page_num + 1
                        }
                    })
        return chunks
    
    def _parse_docx(self, file_path: Path) -> List[Dict]:
        doc = docx.Document(file_path)
        text = "\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        
        return [{
            "text": text,
            "metadata": {
                **self.get_file_metadata(file_path),
                "source_type": "document"
            }
        }]
    
    def _parse_excel(self, file_path: Path) -> List[Dict]:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        chunks = []
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = []
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) for cell in row if cell is not None])
                if row_text.strip():
                    rows.append(row_text)
            
            if rows:
                chunks.append({
                    "text": "\n".join(rows),
                    "metadata": {
                        **self.get_file_metadata(file_path),
                        "source_type": "document",
                        "sheet": sheet_name
                    }
                })
        
        return chunks
    
    def _parse_text(self, file_path: Path) -> List[Dict]:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
        
        return [{
            "text": text,
            "metadata": {
                **self.get_file_metadata(file_path),
                "source_type": "document"
            }
        }]